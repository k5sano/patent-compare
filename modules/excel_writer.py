#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
対比表Excel出力モジュール

テンプレートに準拠した以下のセクションを含むExcelを生成:
1. ヘッダ情報（目的文、案件番号、凡例）
2. 請求項1の対比表
3. 従属請求項の対比表
4. 文献リスト
5. 拒絶理由構成の方針
"""

import html
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers
)
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter

from modules.cited_ref_notation import (
    comment_of as _ref_comment_of,
    display_judgment as _display_judgment,
    expand as _expand_ref,
    normalize as _normalize_ref,
)


# --- スタイル定義 ---
EXCEL_FONT_NAME = "メイリオ"
FONT_TITLE = Font(name=EXCEL_FONT_NAME, size=14, bold=True)
FONT_HEADER = Font(name=EXCEL_FONT_NAME, size=10, bold=True)
FONT_NORMAL = Font(name=EXCEL_FONT_NAME, size=10)
FONT_SMALL = Font(name=EXCEL_FONT_NAME, size=10)
FONT_JUDGMENT = Font(name=EXCEL_FONT_NAME, size=12, bold=True)

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # ○
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # △
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # ×
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_HEADER_LIGHT = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

FONT_WHITE = Font(name=EXCEL_FONT_NAME, size=10, bold=True, color="FFFFFF")

# PDF 注釈と同じキーワードグループ色。Excel はセル内の部分背景が弱いため、
# コメント内では文字色 + 太字で近似する。
GROUP_HIGHLIGHT_COLORS = [
    "FFFF9999", "FFC7A6FF", "FFFF99D9", "FF99BFFF",
    "FF99FFB3", "FFFFD180", "FF80F2D9", "FFBFC4CC",
]
IMPORTANT_TERM_COLOR = "FFEF4444"
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_MAX_EXCEL_CELL_CHARS = 32767

ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGNMENT_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _get_judgment_fill(judgment):
    """判定に応じた背景色を返す"""
    if judgment == "○":
        return FILL_GREEN
    elif judgment == "△":
        return FILL_YELLOW
    elif judgment == "×":
        return FILL_RED
    return None


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    """セルに値とスタイルを設定"""
    if isinstance(value, str):
        value = _excel_safe_text(value)
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _excel_safe_text(value, max_chars=_MAX_EXCEL_CELL_CHARS):
    """Excel XML に安全な文字列へ正規化する。"""
    text = str(value or "")
    if not text:
        return ""
    text = _ILLEGAL_XML_RE.sub("", text)
    if len(text) > max_chars:
        text = text[: max_chars - 15].rstrip() + "\n...[truncated]"
    return text


def _plain_excel_text(value):
    """Excel出力用にHTML/画面表示マーカーをプレーンテキスト化する。"""
    if value is None:
        return ""
    text = str(value)
    if not text:
        return ""
    text = html.unescape(text)
    text = text.replace("<<HL>>", "").replace("<</HL>>", "")
    text = text.replace("<<UL>>", "").replace("<</UL>>", "")
    text = re.sub(r"(?i)<\s*br\s*/?\s*>", "\n", text)
    text = re.sub(r"(?i)</\s*(p|div|li|tr|h[1-6])\s*>", "\n", text)
    text = re.sub(r"(?i)<\s*li(?:\s+[^>]*)?>", "・", text)
    text = re.sub(r"(?i)<\s*/?\s*(ul|ol|table|tbody|thead|tr|td|th|span|strong|b|em|i|p|div|h[1-6])(?:\s+[^>]*)?>", "", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return _excel_safe_text(text.strip())


def _keyword_color(gid):
    try:
        i = int(gid)
    except (TypeError, ValueError):
        i = 1
    return GROUP_HIGHLIGHT_COLORS[(i - 1) % len(GROUP_HIGHLIGHT_COLORS)]


def _rich_font(color, bold=False):
    return InlineFont(rFont=EXCEL_FONT_NAME, color=color, b=bold)


def _keyword_highlight_index(keywords):
    items = []
    for g in keywords or []:
        gid = g.get("group_id")
        color = _keyword_color(gid)
        for kw in g.get("keywords", []) or []:
            term = str((kw or {}).get("term") or "").strip()
            if term:
                items.append({"term": term, "color": color})
    items.sort(key=lambda x: len(x["term"]), reverse=True)
    return items


def _find_keyword_positions(text, index):
    positions = []
    for item in index:
        term = item["term"]
        start = 0
        while True:
            pos = text.find(term, start)
            if pos < 0:
                break
            end = pos + len(term)
            if not any(not (end <= p["start"] or pos >= p["end"]) for p in positions):
                positions.append({"start": pos, "end": end, "color": item["color"]})
            start = end
    positions.sort(key=lambda p: p["start"])
    return positions


def _append_rich_important(rt, text, keywords):
    text = _excel_safe_text(text)
    if not text:
        return
    positions = _find_keyword_positions(text, _keyword_highlight_index(keywords))
    if not positions:
        rt.append(TextBlock(_rich_font(IMPORTANT_TERM_COLOR, bold=True), text))
        return
    prev = 0
    for p in positions:
        if p["start"] > prev:
            rt.append(TextBlock(_rich_font(IMPORTANT_TERM_COLOR, bold=True), text[prev:p["start"]]))
        rt.append(TextBlock(_rich_font(p["color"], bold=True), text[p["start"]:p["end"]]))
        prev = p["end"]
    if prev < len(text):
        rt.append(TextBlock(_rich_font(IMPORTANT_TERM_COLOR, bold=True), text[prev:]))


def _append_rich_keywords(rt, text, keywords):
    text = _excel_safe_text(text)
    if not text:
        return
    positions = _find_keyword_positions(text, _keyword_highlight_index(keywords))
    if not positions:
        rt.append(text)
        return
    prev = 0
    for p in positions:
        if p["start"] > prev:
            rt.append(text[prev:p["start"]])
        rt.append(TextBlock(_rich_font(p["color"], bold=True), text[p["start"]:p["end"]]))
        prev = p["end"]
    if prev < len(text):
        rt.append(text[prev:])


def _build_rich_reason(parts, comment=None, keywords=None):
    rt = CellRichText()
    first = True
    for part in parts or []:
        if not part:
            continue
        if not first:
            rt.append("\n")
        _append_rich_keywords(rt, part, keywords)
        first = False
    if comment:
        if not first:
            rt.append("\n")
        rt.append("（備考: ")
        _append_rich_keywords(rt, comment, keywords)
        rt.append("）")
    return rt if len(rt) else ""


def write_comparison_table(output_path, case_meta, segments, responses,
                           citations_meta=None, keywords=None):
    """対比表Excelを生成 (1 シートのみ)

    Parameters:
        output_path: 出力ファイルパス
        case_meta: 案件メタデータ (case.yaml)
        segments: 請求項分節データ (segments.json)
        responses: {citation_id: response_data} のdict
        citations_meta: {citation_id: citation_json} のdict（任意）
    """
    wb = Workbook()
    _populate_comparison_sheets(wb, case_meta, segments, responses, citations_meta,
                                 main_sheet_title="対比表", drop_default=True,
                                 keywords=keywords)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _populate_comparison_sheets(wb, case_meta, segments, responses, citations_meta=None,
                                 main_sheet_title="対比表", drop_default=False,
                                 keywords=None):
    """対比表 + ペースト用シートを wb に追加する (内部用)。

    drop_default=True なら wb.active が空のデフォルト Sheet (新規 Workbook 直後)
    の場合これを使い回す。False なら create_sheet で新シートを追加する。
    """
    if drop_default and len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1:
        ws = wb.active
        ws.title = main_sheet_title
    else:
        ws = wb.create_sheet(main_sheet_title)

    citations_meta = citations_meta or {}
    case_id = case_meta.get("case_id", "")
    patent_number = case_meta.get("patent_number", case_id)
    patent_title = case_meta.get("patent_title", case_meta.get("title", ""))

    # 引用文献の順序（case.yamlのcitations順）
    citation_order = []
    for cit in case_meta.get("citations", []):
        cid = cit["id"]
        if cid in responses:
            citation_order.append(cid)

    num_citations = len(citation_order)
    if num_citations == 0:
        # 回答がない場合は空の対比表
        citation_order = list(responses.keys())
        num_citations = len(citation_order)

    # 列幅設定
    ws.column_dimensions["A"].width = 6    # 構成要件ID
    ws.column_dimensions["B"].width = 45   # 構成要件テキスト
    for i in range(num_citations):
        col_letter = get_column_letter(3 + i)
        ws.column_dimensions[col_letter].width = 40

    row = 1

    # ===== セクション1: ヘッダ情報 =====
    # 目的文
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=2 + num_citations)
    _set_cell(ws, row, 1,
              "本書は、日本特許法第29条に基づく拒絶理由の構成を目的とした "
              "先行技術文献との対比表です。出願日以前の公知文献のみを対象とします。",
              font=FONT_SMALL, alignment=ALIGNMENT_LEFT_CENTER)
    row += 1

    # 案件情報
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=2 + num_citations)
    _set_cell(ws, row, 1, f"{patent_number}　{patent_title}　請求項分節・対比表",
              font=FONT_TITLE, alignment=ALIGNMENT_LEFT_CENTER)
    row += 2

    # 凡例
    _set_cell(ws, row, 1, "凡例:", font=FONT_HEADER)
    _set_cell(ws, row, 2, "○＝一致（同一又は実質同一の構成が開示）", font=FONT_NORMAL,
              fill=FILL_GREEN)
    row += 1
    _set_cell(ws, row, 2, "△＝部分一致（上位概念での開示、数値範囲の一部重複等）",
              font=FONT_NORMAL, fill=FILL_YELLOW)
    row += 1
    _set_cell(ws, row, 2, "×＝不一致（対応する記載なし）", font=FONT_NORMAL,
              fill=FILL_RED)
    row += 2

    # ===== セクション2: 請求項順の対比表 =====
    # 独立請求項は分節単位、従属請求項は請求項単位で、segments.json の順番を保つ。
    for idx, claim in enumerate(segments):
        if idx > 0:
            row += 1
        if claim.get("is_independent") or claim.get("claim_number") == 1:
            row = _write_claim_comparison(ws, row, claim, citation_order,
                                          responses, citations_meta, case_meta,
                                          keywords=keywords)
        else:
            row = _write_sub_claims_table(ws, row, [claim], citation_order,
                                          responses, citations_meta, keywords=keywords)
        row += 2

    # ===== セクション4: 文献リスト =====
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=2 + num_citations)
    _set_cell(ws, row, 1, "拒絶理由構成のための先行技術文献調査結果",
              font=FONT_TITLE, fill=FILL_SECTION)
    row += 1

    row = _write_document_list(ws, row, citation_order, responses,
                                citations_meta, case_meta, num_citations,
                                keywords=keywords)
    row += 2

    # ===== セクション5: 拒絶理由構成の方針 =====
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=2 + num_citations)
    _set_cell(ws, row, 1, "拒絶理由構成の方針（分析コメント）",
              font=FONT_TITLE, fill=FILL_SECTION)
    row += 1

    row = _write_rejection_strategy(ws, row, citation_order, responses, num_citations)

    # ===== シート2: ペースト用（既存対比表への貼付用、コンパクト記法） =====
    _write_paste_sheet(wb, segments, citation_order, responses, citations_meta, case_meta)


def _strip_comment_memo_from_loc(raw):
    """``cited_location`` から `"...` コメントと `//...` メモを除去し、参照記法のみを返す。

    JS 側 `_stripCommentMemoFromLoc` と同等のロジック。
    """
    if not raw:
        return ""
    out = []
    for tok in str(raw).split(";"):
        tok = tok.strip()
        if not tok or tok.startswith('"') or tok.startswith("/") or tok.startswith("//"):
            continue
        # トークン内の " / // 以降を切り捨て
        cuts = []
        i = tok.find('"')
        if i >= 0:
            cuts.append(i)
        i = tok.find("/")
        if i >= 0:
            cuts.append(i)
        i = tok.find("//")
        if i >= 0:
            cuts.append(i)
        if cuts:
            tok = tok[:min(cuts)].strip()
        if tok:
            out.append(tok)
    return ";".join(out)


def _short_reason(s, limit=None):
    """judgment_reason を最初の文 (「。」まで) で切る。途中切り捨て (…) はしない。

    LLM プロンプト側で「相違点を 1 文で簡潔に」と指示している前提なので、
    無理な文字数制限は掛けず、最初の句点までを返す。複数文ある場合のみ
    末尾「。」を 1 個分だけ落とす。
    """
    if not s:
        return ""
    import re
    t = re.sub(r"[\r\n]+", " ", str(s))
    t = re.sub(r"\s{2,}", " ", t).strip()
    m = re.search(r"[。．]", t)
    if m and m.start() > 0:
        t = t[: m.start()]
    return t


def _format_comp_for_paste(comp):
    """comparison/sub_claim 1 件分を「貼付用」セル内容に整形。

    JS 側 `_formatCompForPaste` と同等。
      - judgment ○ → prefix なし
      - judgment △ → "?"
      - judgment × → "!"
      - cited_location は raw 記法 (展開しない)。" コメントと // メモは除外。
      - 末尾: コメント (manual) があれば "/<comment>"。
        なければ judgment_reason (△/× のみ、短縮形) を "/" で付ける。
    """
    if not isinstance(comp, dict):
        return ""
    j = (comp.get("judgment") or "").strip()
    prefix = ""
    if j == "△":
        prefix = "?"
    elif j == "×":
        prefix = "!"

    raw = comp.get("cited_location") or ""
    loc_only = _strip_comment_memo_from_loc(raw)
    if loc_only:
        try:
            loc_only = _normalize_ref(loc_only)
        except Exception:
            pass

    # manual comment ("...") 抽出
    from modules.cited_ref_notation import comment_of
    comment = (comment_of(raw) or "").strip()
    if not comment and j in ("△", "×") and comp.get("judgment_reason"):
        comment = _short_reason(comp.get("judgment_reason"))

    out = prefix + loc_only
    if comment:
        # タブ/改行は貼付時にセルが暴れるので空白に置換
        clean = comment.replace("\t", " ").replace("\r", " ").replace("\n", " ").strip()
        out += "/" + clean
    return out


def _claim_full_text(claim):
    """請求項 dict から全文を組み立てる。"""
    if not claim:
        return ""
    if claim.get("full_text"):
        return claim["full_text"]
    parts = [s.get("text", "") for s in (claim.get("segments") or []) if s.get("text")]
    return "".join(parts)


def _write_paste_sheet(wb, segments, citation_order, responses, citations_meta, case_meta):
    """既存対比表への貼付専用シート。

    各引例ごとに 1 列。1 行 = 1 分節 (請求項1) または 1 従属請求項。
    請求項1 の分節は連続、請求項2 以降の各従属請求項の前に空行 1 つ。

    レイアウト:
      | 構成要件ID | 構成要件テキスト | 文献1 | 文献2 | ... |
      | 1a        | (claim1 seg text) | CL1,..| ...   | ... |
      | 1b        | ...               | ...   | ...   | ... |
      | (empty row before claim 2)                          |
      | 2a        | (claim 2 text)    | 13,51 | ...   | ... |
      | ...                                                  |
    """
    ws = wb.create_sheet("ペースト用")
    num_cit = len(citation_order)

    # 列幅
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    for i in range(num_cit):
        ws.column_dimensions[get_column_letter(3 + i)].width = 30

    # ヘッダ行
    row = 1
    _set_cell(ws, row, 1, "ID", font=FONT_WHITE, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    _set_cell(ws, row, 2, "構成要件", font=FONT_WHITE, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    for i, cit_id in enumerate(citation_order):
        cit_info = None
        for c in case_meta.get("citations", []):
            if c["id"] == cit_id:
                cit_info = c
                break
        label = cit_info["label"] if cit_info else cit_id
        _set_cell(ws, row, 3 + i, f"{label}\n({cit_id})",
                  font=FONT_WHITE, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    row += 1

    # 案内行
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + num_cit)
    _set_cell(ws, row, 1,
              "貼付用フォーマット: 各引例の列をコピーして既存対比表へ縦方向に貼付け。"
              "判定 ○ は prefix なし、△ は ?、× は ! を冠。/ 以降はコメント。",
              font=FONT_SMALL, alignment=ALIGNMENT_LEFT,
              fill=FILL_HEADER_LIGHT, border=THIN_BORDER)
    row += 1

    # segments.json の請求項順を保つ。独立請求項は分節単位、従属請求項は請求項単位。
    for idx_claim, claim in enumerate(segments):
        if idx_claim > 0:
            for col in range(1, 3 + num_cit):
                _set_cell(ws, row, col, "", border=THIN_BORDER)
            row += 1
        claim_num = claim.get("claim_number")
        if claim.get("is_independent") or claim_num == 1:
            for seg in claim.get("segments", []):
                seg_id = seg.get("id", "")
                seg_text = seg.get("text", "")
                _set_cell(ws, row, 1, seg_id, font=FONT_NORMAL,
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
                _set_cell(ws, row, 2, seg_text, font=FONT_NORMAL,
                          alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
                for i, cit_id in enumerate(citation_order):
                    resp = responses.get(cit_id, {})
                    comp = _find_comparison(resp, seg_id)
                    cell_val = _format_comp_for_paste(comp) if comp else ""
                    fill = _get_judgment_fill(comp.get("judgment", "")) if comp else None
                    _set_cell(ws, row, 3 + i, cell_val, font=FONT_NORMAL,
                              alignment=ALIGNMENT_LEFT, border=THIN_BORDER, fill=fill)
                row += 1
            continue

        # 通番ID: "{claim_num}a" 形 (1a/1b と並ぶ慣行に合わせる)
        # 元データに segments[0].id があればそれを優先 (既存形式へ追従)
        seg_id_default = f"{claim_num}a"
        first_seg_id = ""
        for s in (claim.get("segments") or []):
            if s.get("id"):
                first_seg_id = s["id"]
                break
        seg_id = first_seg_id or seg_id_default
        full_text = _claim_full_text(claim)

        _set_cell(ws, row, 1, seg_id, font=FONT_NORMAL,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        _set_cell(ws, row, 2, full_text, font=FONT_NORMAL,
                  alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
        for i, cit_id in enumerate(citation_order):
            resp = responses.get(cit_id, {})
            sub = _find_sub_claim(resp, claim_num)
            cell_val = _format_comp_for_paste(sub) if sub else ""
            fill = _get_judgment_fill(sub.get("judgment", "")) if sub else None
            _set_cell(ws, row, 3 + i, cell_val, font=FONT_NORMAL,
                      alignment=ALIGNMENT_LEFT, border=THIN_BORDER, fill=fill)
        row += 1

    # 列幅自動調整は openpyxl 側で限定的なので固定値で運用


def _write_claim_comparison(ws, row, claim, citation_order, responses,
                            citations_meta, case_meta, keywords=None):
    """請求項の対比表を書き込む"""
    num_citations = len(citation_order)

    # ヘッダ行: 構成要件 | 請求項X | 文献1 | 文献2 | ...
    _set_cell(ws, row, 1, "ID", font=FONT_WHITE, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    _set_cell(ws, row, 2, f"請求項{claim['claim_number']} 構成要件",
              font=FONT_WHITE, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)

    for i, cit_id in enumerate(citation_order):
        cit_info = None
        for c in case_meta.get("citations", []):
            if c["id"] == cit_id:
                cit_info = c
                break
        label = cit_info["label"] if cit_info else cit_id
        role = cit_info.get("role", "") if cit_info else ""
        # response から category_suggestion を読んで X/Y/A バッジを文献名行に付与
        category = ""
        resp = responses.get(cit_id, {}) or {}
        cat_raw = (resp.get("category_suggestion") or "").strip().upper()
        if cat_raw and cat_raw[:1] in ("X", "Y", "A"):
            category = cat_raw[:1]
        prefix = f"[{category}] " if category else ""
        header_text = f"{prefix}{label}\n({cit_id})\n{role}"
        _set_cell(ws, row, 3 + i, header_text,
                  font=FONT_WHITE, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    row += 1

    # 各構成要件
    for seg in claim["segments"]:
        seg_id = seg["id"]
        seg_text = seg["text"]

        # 判定行
        _set_cell(ws, row, 1, seg_id, font=FONT_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER,
                  fill=FILL_HEADER_LIGHT)
        _set_cell(ws, row, 2, seg_text, font=FONT_NORMAL,
                  alignment=ALIGNMENT_LEFT, border=THIN_BORDER)

        for i, cit_id in enumerate(citation_order):
            resp = responses.get(cit_id, {})
            comp = _find_comparison(resp, seg_id)
            if comp:
                judgment_raw = comp.get("judgment", "")
                fill = _get_judgment_fill(judgment_raw)
                # ○ は「先頭に何もつけない」慣行 → 表示は空、塗りは緑のまま
                judgment_disp = _display_judgment(judgment_raw)
                _set_cell(ws, row, 3 + i, judgment_disp,
                          font=FONT_JUDGMENT, fill=fill,
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
            else:
                _set_cell(ws, row, 3 + i, "-",
                          alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        row += 1

        # 理由行: cited_location は記法を展開して書き込む。
        #   コメント部分 ("...) は備考扱いとして判定理由の末尾に括弧書きで合成。
        #   メモ部分 (//...) は対外出力なので含めない。
        _set_cell(ws, row, 1, "", border=THIN_BORDER)
        _set_cell(ws, row, 2, "", border=THIN_BORDER)

        for i, cit_id in enumerate(citation_order):
            resp = responses.get(cit_id, {})
            comp = _find_comparison(resp, seg_id)
            if comp:
                reason_parts = []
                if comp.get("judgment_reason"):
                    reason_parts.append(comp["judgment_reason"])
                comment = ""
                cited_loc_raw = comp.get("cited_location", "")
                if cited_loc_raw:
                    expanded = _expand_ref(cited_loc_raw, with_comment=False)
                    if expanded:
                        reason_parts.append(f"[{expanded}]")
                    comment = _ref_comment_of(cited_loc_raw)
                if comp.get("cited_text"):
                    reason_parts.append(f"「{comp['cited_text'][:100]}」")
                reason_text = _build_rich_reason(reason_parts, comment=comment, keywords=keywords)
                _set_cell(ws, row, 3 + i, reason_text,
                          font=FONT_SMALL, alignment=ALIGNMENT_LEFT,
                          border=THIN_BORDER)
            else:
                _set_cell(ws, row, 3 + i, "", border=THIN_BORDER)
        row += 1

    return row


def _write_sub_claims_table(ws, row, sub_claims, citation_order, responses,
                            citations_meta, keywords=None):
    """従属請求項の対比表"""
    num_citations = len(citation_order)

    # ヘッダ
    _set_cell(ws, row, 1, "請求項", font=FONT_WHITE, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    _set_cell(ws, row, 2, "追加限定事項", font=FONT_WHITE, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    for i, cit_id in enumerate(citation_order):
        _set_cell(ws, row, 3 + i, cit_id, font=FONT_WHITE, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    row += 1

    for claim in sub_claims:
        claim_num = claim["claim_number"]

        # 各分節を表示
        for seg in claim["segments"]:
            _set_cell(ws, row, 1, f"請求項{claim_num}\n{seg['id']}",
                      font=FONT_HEADER, alignment=ALIGNMENT_CENTER,
                      border=THIN_BORDER, fill=FILL_HEADER_LIGHT)
            _set_cell(ws, row, 2, seg["text"], font=FONT_NORMAL,
                      alignment=ALIGNMENT_LEFT, border=THIN_BORDER)

            for i, cit_id in enumerate(citation_order):
                resp = responses.get(cit_id, {})
                # sub_claimsから検索
                sub_comp = _find_sub_claim(resp, claim_num)
                if sub_comp:
                    judgment_raw = sub_comp.get("judgment", "")
                    fill = _get_judgment_fill(judgment_raw)
                    judgment_disp = _display_judgment(judgment_raw)
                    parts = []
                    if judgment_disp:
                        parts.append(judgment_disp)
                    if sub_comp.get("judgment_reason"):
                        parts.append(sub_comp["judgment_reason"])
                    comment = ""
                    cited_loc_raw = sub_comp.get("cited_location", "")
                    if cited_loc_raw:
                        expanded = _expand_ref(cited_loc_raw, with_comment=False)
                        if expanded:
                            parts.append(f"[{expanded}]")
                        comment = _ref_comment_of(cited_loc_raw)
                    cell_text = _build_rich_reason(parts, comment=comment, keywords=keywords)
                    _set_cell(ws, row, 3 + i, cell_text,
                              font=FONT_NORMAL, fill=fill,
                              alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
                else:
                    _set_cell(ws, row, 3 + i, "-",
                              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
            row += 1

    return row


def _write_document_list(ws, row, citation_order, responses, citations_meta,
                         case_meta, num_citations, keywords=None):
    """文献リストセクション"""
    # ヘッダ
    headers = ["No", "文献名・タイトル", "概要・拒絶理由との関連性"]
    cols = min(len(headers), 2 + num_citations)

    _set_cell(ws, row, 1, "No", font=FONT_WHITE, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2)
    _set_cell(ws, row, 2, "文献名・タイトル", font=FONT_WHITE, fill=FILL_HEADER,
              alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    if num_citations > 0:
        ws.merge_cells(start_row=row, start_column=3,
                       end_row=row, end_column=2 + num_citations)
        _set_cell(ws, row, 3, "概要・拒絶理由との関連性",
                  font=FONT_WHITE, fill=FILL_HEADER,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
    row += 1

    for idx, cit_id in enumerate(citation_order, 1):
        resp = responses.get(cit_id, {})
        cit_info = None
        for c in case_meta.get("citations", []):
            if c["id"] == cit_id:
                cit_info = c
                break

        label = cit_info["label"] if cit_info else f"文献{idx}"
        role = cit_info.get("role", "") if cit_info else ""
        category = resp.get("category_suggestion", "")
        summary = resp.get("overall_summary", "")
        relevance = resp.get("rejection_relevance", "")

        _set_cell(ws, row, 1, idx, font=FONT_NORMAL,
                  alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
        _set_cell(ws, row, 2, f"{label}\n({cit_id})\n役割: {role}\nカテゴリ: {category}",
                  font=FONT_NORMAL, alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
        if num_citations > 0:
            ws.merge_cells(start_row=row, start_column=3,
                           end_row=row, end_column=2 + num_citations)
            _set_cell(ws, row, 3, _build_rich_keyword_text(f"{summary}\n\n{relevance}", keywords),
                      font=FONT_NORMAL, alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
        ws.row_dimensions[row].height = 60
        row += 1

    return row


def _write_rejection_strategy(ws, row, citation_order, responses, num_citations):
    """拒絶理由構成の方針セクション"""
    total_cols = 2 + num_citations

    # 自動的に方針案を生成
    strategies = []

    # 各文献のカテゴリを確認
    x_docs = []  # 単独拒絶候補
    y_docs = []  # 組合せ拒絶候補
    for cit_id in citation_order:
        resp = responses.get(cit_id, {})
        cat = resp.get("category_suggestion", "")
        if cat == "X":
            x_docs.append(cit_id)
        elif cat == "Y":
            y_docs.append(cit_id)

    strategy_num = 1
    if x_docs:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=total_cols)
        _set_cell(ws, row, 1,
                  f"方針{strategy_num}: 29条2項（進歩性）— "
                  f"{', '.join(x_docs)} を主引例として単独で拒絶理由を構成",
                  font=FONT_NORMAL, alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
        row += 1
        strategy_num += 1

    if len(y_docs) >= 2:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=total_cols)
        _set_cell(ws, row, 1,
                  f"方針{strategy_num}: 29条2項（進歩性）— "
                  f"{', '.join(y_docs)} の組合せで拒絶理由を構成",
                  font=FONT_NORMAL, alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
        row += 1
        strategy_num += 1

    if not x_docs and not y_docs:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=total_cols)
        _set_cell(ws, row, 1,
                  "（自動分析結果: 現在の文献では拒絶理由の構成が困難な可能性があります。"
                  "追加の文献調査を検討してください。）",
                  font=FONT_NORMAL, alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
        row += 1

    # 留意点
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=total_cols)
    _set_cell(ws, row, 1, "留意点:", font=FONT_HEADER, alignment=ALIGNMENT_LEFT)
    row += 1
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=total_cols)
    _set_cell(ws, row, 1,
              "- △判定の構成要件は、審査官の判断により○に引き上げられる可能性があります\n"
              "- ×判定の構成要件については、副引例や技術常識で補う必要があります\n"
              "- 上記は自動分析結果です。最終判断はサーチャー・審査官が行ってください",
              font=FONT_SMALL, alignment=ALIGNMENT_LEFT)
    row += 1

    return row


def _find_comparison(resp, segment_id):
    """回答データから指定された構成要件の比較結果を検索。

    旧プロンプトでは、請求項2/4/8/9のような後続独立請求項が
    ``sub_claims`` に入っていた。正式には再対比対象だが、完成版Excelの
    空欄を減らすため、``requirement_id`` が一致する旧データを暫定利用する。
    """
    for comp in resp.get("comparisons", []):
        if comp.get("requirement_id") == segment_id:
            return comp
    for sc in resp.get("sub_claims", []):
        if sc.get("requirement_id") == segment_id:
            return sc
    m = re.match(r"^(\d+)", str(segment_id or ""))
    claim_number = int(m.group(1)) if m else None
    if claim_number is not None:
        for sc in resp.get("sub_claims", []):
            if sc.get("claim_number") == claim_number:
                return sc
    return None


def _find_sub_claim(resp, claim_number):
    """回答データから従属請求項の判定を検索"""
    for sc in resp.get("sub_claims", []):
        if sc.get("claim_number") == claim_number:
            return sc
    return None


# ============================================================
# 完成版対比表 (3 タブ統合) writer
# ============================================================

def write_full_report(output_path, case_meta, segments, responses,
                      citations_meta=None, hongan_analysis=None, inventive_step=None,
                      keywords=None):
    """3 タブ統合 Excel を生成。

    タブ構成:
        - 本願解析結果   (hongan_analysis_v0.1.yaml の構造化結果)
        - 対比表         (write_comparison_table と同じ + ペースト用シート)
        - 進歩性判断     (inventive_step.json のパース結果)
    """
    wb = Workbook()
    # デフォルト Sheet を本願解析結果に転用 (空なら使い回し)
    _populate_hongan_analysis_sheet(wb, case_meta, hongan_analysis, drop_default=True)
    _populate_comparison_sheets(wb, case_meta, segments, responses, citations_meta,
                                 main_sheet_title="対比表", drop_default=False,
                                 keywords=keywords)
    _populate_inventive_step_sheet(wb, case_meta, inventive_step, keywords=keywords)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _ensure_sheet(wb, title, drop_default=False):
    """新規シート作成 (drop_default なら 空のデフォルト Sheet を使い回す)。"""
    if drop_default and len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1:
        ws = wb.active
        ws.title = title
        return ws
    return wb.create_sheet(title)


def _populate_hongan_analysis_sheet(wb, case_meta, hongan_analysis, drop_default=False):
    """本願解析結果 (analysis/hongan_analysis.json) をシートに展開。"""
    ws = _ensure_sheet(wb, "本願解析結果", drop_default=drop_default)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 90

    case_id = case_meta.get("case_id", "")
    title = case_meta.get("patent_title") or case_meta.get("title", "")
    pn = case_meta.get("patent_number", case_id)

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set_cell(ws, row, 1, f"{pn}　{title}　本願解析結果",
              font=FONT_TITLE, alignment=ALIGNMENT_LEFT_CENTER)
    row += 1

    if not hongan_analysis or not hongan_analysis.get("sections"):
        _set_cell(ws, row, 1, "(本願分析が未実行です。Step 2 SUB 3 で実行してください)",
                  font=FONT_SMALL, alignment=ALIGNMENT_LEFT)
        return

    for sec in hongan_analysis.get("sections") or []:
        # セクション見出し
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        _set_cell(ws, row, 1, f"{sec.get('id')} {sec.get('title', '')}",
                  font=FONT_HEADER, fill=FILL_HEADER_LIGHT, alignment=ALIGNMENT_LEFT_CENTER)
        row += 1

        for it in sec.get("items") or []:
            iid = it.get("id", "")
            label = it.get("label", "")
            value = it.get("value")
            value_text = _format_analysis_value(value)
            _set_cell(ws, row, 1, iid, font=FONT_NORMAL,
                      alignment=ALIGNMENT_CENTER, border=THIN_BORDER)
            _set_cell(ws, row, 2, label, font=FONT_NORMAL,
                      alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
            _set_cell(ws, row, 3, value_text, font=FONT_NORMAL,
                      alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
            row += 1
        row += 1


def _format_analysis_value(value):
    """本願分析の value (string / list / dict / None) を Excel 表示用に変換。

    空値処理:
      - None / "" / [] / {} → "(未取得)"
      - dict 内の空値は当該キーをスキップ (キーだけ残らないように)
      - "_" 始まりの内部キー (_note 等) はスキップ
    """
    if value is None or value == "":
        return "(未取得)"
    if isinstance(value, str):
        return _plain_excel_text(value) or "(未取得)"
    if isinstance(value, list):
        if not value:
            return "(未取得)"
        if all(isinstance(x, str) for x in value):
            joined = "、".join(_plain_excel_text(x) for x in value if _plain_excel_text(x))
            return joined or "(未取得)"
        lines = []
        for x in value:
            if isinstance(x, dict):
                code = (x.get("code") or "").strip()
                lab = _plain_excel_text(x.get("label") or "")
                if code or lab:
                    lines.append(f"{code}（{lab}）" if lab else code)
                    continue
                formatted = _format_analysis_value(x)
                if formatted and formatted != "(未取得)":
                    lines.append(formatted)
            elif x not in (None, ""):
                lines.append(_plain_excel_text(x))
        return "\n".join(lines) if lines else "(未取得)"
    if isinstance(value, dict):
        parts = []
        for k, v in value.items():
            if isinstance(k, str) and k.startswith("_"):
                continue  # _note 等の内部メタはスキップ
            clean_key = _plain_excel_text(k)
            if isinstance(v, dict) and "items" in v:
                # F-term grouped: {theme: {theme_label, items: [{code, label}, ...]}}
                items = v.get("items") or []
                joined = "、".join(
                    f"{x.get('code', '')}（{_plain_excel_text(x.get('label', ''))}）" if x.get("label")
                    else x.get("code", "")
                    for x in items
                    if x.get("code") or x.get("label")
                )
                if not joined:
                    continue
                tlab = _plain_excel_text(v.get("theme_label") or "")
                parts.append(f"{clean_key}（{tlab}）: {joined}" if tlab else f"{clean_key}: {joined}")
            elif isinstance(v, list):
                if not v:
                    continue
                if all(isinstance(x, dict) for x in v):
                    formatted = "、".join(
                        f"{x.get('code', '')}（{_plain_excel_text(x.get('label', ''))}）" if x.get("label")
                        else x.get("code", "")
                        for x in v
                        if x.get("code") or x.get("label")
                    )
                    if not formatted:
                        formatted = "\n".join(
                            _format_analysis_value(x)
                            for x in v
                            if _format_analysis_value(x) != "(未取得)"
                        )
                else:
                    formatted = "、".join(_plain_excel_text(x) for x in v if x not in (None, ""))
                if formatted:
                    parts.append(f"{clean_key}: {formatted}")
            elif isinstance(v, dict):
                # ネスト dict は再帰的に整形
                formatted = _format_analysis_value(v)
                if formatted and formatted != "(未取得)":
                    parts.append(f"{clean_key}: {formatted}")
            elif v not in (None, ""):
                parts.append(f"{clean_key}: {_plain_excel_text(v)}")
        return "\n".join(parts) if parts else "(未取得)"
    return _plain_excel_text(value)


def _build_rich_keyword_text(value, keywords=None):
    text = _format_analysis_value(value)
    if not text or text == "(未取得)":
        return text
    index = _keyword_highlight_index(keywords)
    if not _find_keyword_positions(text, index):
        return text
    rt = CellRichText()
    _append_rich_keywords(rt, text, keywords)
    return rt


def _populate_inventive_step_sheet(wb, case_meta, inventive_step, keywords=None):
    """進歩性判断 (inventive_step.json) をシートに展開。

    実構造 (modules.inventive_step_analyzer 由来):
      - primary_reference: {document_id, selection_reason}
      - common_features: [{description, segment_ids}]
      - differences: [{segment_id, description, technical_significance, resolution: {...}}]
      - advantageous_effects: {claimed_effects, assessment, ...}
      - overall_assessment: {inventive_step, reasoning, rejection_logic,
                             vulnerable_points, strengthening_suggestions}
    """
    ws = _ensure_sheet(wb, "進歩性判断")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100

    case_id = case_meta.get("case_id", "")
    title = case_meta.get("patent_title") or case_meta.get("title", "")
    pn = case_meta.get("patent_number", case_id)

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    _set_cell(ws, row, 1, f"{pn}　{title}　進歩性判断 (JPO 審査基準ベース)",
              font=FONT_TITLE, alignment=ALIGNMENT_LEFT_CENTER)
    row += 2

    if not inventive_step or not isinstance(inventive_step, dict):
        _set_cell(ws, row, 1, "(進歩性判断が未実行です。Step 6 で実行してください)",
                  font=FONT_SMALL, alignment=ALIGNMENT_LEFT)
        return

    def _section(label):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _set_cell(ws, row, 1, label, font=FONT_HEADER,
                  fill=FILL_HEADER_LIGHT, alignment=ALIGNMENT_LEFT_CENTER)
        row += 1

    def _kv(label, value):
        nonlocal row
        if value is None or value == "":
            return
        _set_cell(ws, row, 1, label, font=FONT_HEADER,
                  alignment=ALIGNMENT_LEFT_CENTER, border=THIN_BORDER)
        _set_cell(ws, row, 2, _build_rich_keyword_text(value, keywords),
                  font=FONT_NORMAL,
                  alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
        row += 1

    # ===== 総合評価 (最重要) =====
    oa = inventive_step.get("overall_assessment") or {}
    if oa:
        _section("【総合評価】")
        _kv("進歩性", oa.get("inventive_step"))
        _kv("評価理由", oa.get("reasoning"))
        _kv("拒絶理由の論理構成", oa.get("rejection_logic"))
        _kv("反論されやすいポイント", oa.get("vulnerable_points"))
        _kv("論理強化の提案", oa.get("strengthening_suggestions"))
        row += 1

    # ===== 主引用発明 =====
    pr = inventive_step.get("primary_reference") or {}
    if pr:
        _section("【主引用発明】")
        _kv("文献ID", pr.get("document_id"))
        _kv("選定理由", pr.get("selection_reason"))
        row += 1

    # ===== 一致点 =====
    cf = inventive_step.get("common_features") or []
    if cf:
        _section(f"【一致点 ({len(cf)} 件)】")
        for i, item in enumerate(cf, 1):
            seg_ids = "、".join(item.get("segment_ids") or []) or "-"
            _set_cell(ws, row, 1, f"#{i} 構成要件 {seg_ids}", font=FONT_HEADER,
                      alignment=ALIGNMENT_LEFT_CENTER, border=THIN_BORDER)
            _set_cell(ws, row, 2,
                      _build_rich_keyword_text(item.get("description", ""), keywords),
                      font=FONT_NORMAL,
                      alignment=ALIGNMENT_LEFT, border=THIN_BORDER)
            row += 1
        row += 1

    # ===== 相違点 =====
    diffs = inventive_step.get("differences") or []
    if diffs:
        _section(f"【相違点 ({len(diffs)} 件)】")
        for i, d in enumerate(diffs, 1):
            seg = d.get("segment_id", "")
            _kv(f"#{i} {seg} 相違点", d.get("description"))
            _kv(f"#{i} {seg} 技術的意義", d.get("technical_significance"))
            res = d.get("resolution") or {}
            if res:
                _kv(f"#{i} {seg} 解決方法", res.get("method"))
                _kv(f"#{i} {seg} 副引例", res.get("secondary_reference"))
                _kv(f"#{i} {seg} 設計変更類型", res.get("design_change_type"))
                _kv(f"#{i} {seg} 結論", res.get("conclusion"))
                mot = res.get("motivation") or {}
                if mot:
                    _kv(f"#{i} {seg} 動機: 技術分野", mot.get("technical_field"))
                    _kv(f"#{i} {seg} 動機: 課題の共通性", mot.get("common_problem"))
                    _kv(f"#{i} {seg} 動機: 機能の共通性", mot.get("common_function"))
                    _kv(f"#{i} {seg} 動機: 示唆", mot.get("suggestion"))
                inhibit = res.get("inhibiting_factors") or []
                if inhibit:
                    joined = "\n".join(f"・{x}" for x in inhibit)
                    _kv(f"#{i} {seg} 阻害要因", joined)
            row += 1

    # ===== 有利な効果 =====
    ae = inventive_step.get("advantageous_effects") or {}
    if ae:
        _section("【有利な効果】")
        _kv("主張する効果", ae.get("claimed_effects"))
        _kv("効果の評価", ae.get("assessment"))
        # bool フラグ群
        for k, label in [
            ("is_heterogeneous", "異質効果"),
            ("is_predictable", "予測可能性"),
            ("is_remarkably_superior", "顕著な効果"),
        ]:
            if k in ae:
                _kv(label, "あり" if ae[k] else "なし")
