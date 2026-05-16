"""comparison_service の軽量スモークテスト。"""
from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

from modules.excel_writer import _find_comparison
from services.comparison_service import (
    check_segments_freshness,
    export_excel,
    export_full_report,
    generate_prompt_single,
    save_response_single,
)


SNAPSHOTS = Path(__file__).resolve().parent / "fixtures" / "snapshots"


def _prompt_sections_snapshot(prompt):
    lines = []
    for line in prompt.splitlines():
        s = line.rstrip()
        st = s.strip()
        if (
            s.startswith("## ")
            or s.startswith("### 請求項")
            or s.startswith("### グループ")
            or s.startswith("- **")
            or s.startswith("- 成分")
            or s.startswith("【請求項")
            or s.startswith("【001")
            or st.startswith('"document_id"')
            or st.startswith('"requirement_id"')
            or "全ての構成要件" in s
            or "全ての独立請求項の構成要件" in s
        ):
            lines.append(s)
    return "\n".join(lines)


def test_generate_prompt_single_contains_required_sections(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")

    result, code = generate_prompt_single("smoke", "JP2030000002A")

    assert code == 200
    assert result["char_count"] > 1000
    prompt = result["prompt"]
    assert "## タスク" in prompt
    assert "## 本願の請求項 構成要件" in prompt
    assert "JP2030000002A" in prompt
    assert (case_dir / "prompts" / "JP2030000002A_prompt.txt").exists()


def test_generate_prompt_single_sections_match_snapshot(copy_case_fixture):
    copy_case_fixture("smoke")
    result, code = generate_prompt_single("smoke", "JP2030000002A")

    assert code == 200
    actual = _prompt_sections_snapshot(result["prompt"])
    expected = (SNAPSHOTS / "generate_prompt_single_smoke.sections.txt").read_text(encoding="utf-8").rstrip()
    assert actual == expected


def test_save_response_single_writes_snapshot_json(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    raw = json.dumps({
        "document_id": "JP2030000002A",
        "comparisons": [
            {
                "requirement_id": "1A",
                "judgment": "○",
                "cited_location": "0010;0011",
                "judgment_reason": "成分Aが記載されている。",
            },
            {
                "requirement_id": "1B",
                "judgment": "△",
                "cited_location": "T1;T2",
                "judgment_reason": "成分Bは表で一部記載されている。",
            },
        ],
        "overall_summary": "保存テスト",
        "category_suggestion": "Y",
    }, ensure_ascii=False)

    result, code = save_response_single("smoke", "JP2030000002A", raw)

    assert code == 200
    assert result["success"] is True
    saved = (case_dir / "responses" / "JP2030000002A.json").read_text(encoding="utf-8").rstrip()
    expected = (SNAPSHOTS / "save_response_single_smoke.json").read_text(encoding="utf-8").rstrip()
    assert saved == expected


def test_excel_comparison_lookup_uses_legacy_sub_claim_requirement_id():
    resp = {
        "comparisons": [{"requirement_id": "1A", "judgment": "○"}],
        "sub_claims": [{"claim_number": 9, "requirement_id": "9A", "judgment": "△"}],
    }

    assert _find_comparison(resp, "1A")["judgment"] == "○"
    assert _find_comparison(resp, "9A")["judgment"] == "△"
    assert _find_comparison(resp, "9B")["judgment"] == "△"


def test_export_excel_writes_workbook(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")

    result, code = export_excel("smoke")

    assert code == 200
    assert result["success"] is True
    out = Path(result["path"])
    assert out.exists()
    assert out.parent == case_dir / "output"
    wb = load_workbook(out)
    assert "対比表" in wb.sheetnames
    ws = wb["対比表"]
    values = [cell.value for row in ws.iter_rows(max_row=12, max_col=4) for cell in row]
    assert "JP2030000002A" in " ".join(str(v or "") for v in values)


def test_export_excel_expected_cells_are_stable(copy_case_fixture):
    copy_case_fixture("smoke")
    result, code = export_excel("smoke")

    assert code == 200
    ws = load_workbook(result["path"])["対比表"]
    assert ws["A2"].value == "特開2030-000001　テスト組成物　請求項分節・対比表"
    assert ws["A8"].value == "ID"
    assert ws["B8"].value == "請求項1 構成要件"
    assert ws["C8"].value == "[X] JP2030000002A\n(JP2030000002A)\n主引例"
    assert ws["A9"].value == "1A"
    assert ws["B9"].value == "成分Aを含む"
    assert ws["C10"].value == "段落0010に成分Aが記載されている。\n[段落【0010】]"
    assert ws["A11"].value == "1B"
    assert ws["B11"].value == "成分Bを含む"
    assert ws["C12"].value == "段落0011に成分Bが記載されている。\n[段落【0011】]"


def test_export_excel_highlights_comment_terms(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    resp_path = case_dir / "responses" / "JP2030000002A.json"
    response = json.loads(resp_path.read_text(encoding="utf-8"))
    response["comparisons"][0]["cited_location"] = "10;/成分Aと未知語"
    resp_path.write_text(json.dumps(response, ensure_ascii=False, indent=2), encoding="utf-8")

    result, code = export_excel("smoke")

    assert code == 200
    ws = load_workbook(result["path"], rich_text=True)["対比表"]
    value = ws["C10"].value
    assert isinstance(value, CellRichText)
    assert "（備考: 成分Aと未知語）" in str(value)
    blocks = [b for b in value if isinstance(b, TextBlock)]
    assert any(b.text == "成分A" and b.font.color.rgb == "FFFF9999" for b in blocks)
    assert any(b.text == "成分A" and b.font.rFont == "メイリオ" for b in blocks)
    assert not any("未知語" in b.text and b.font.color.rgb == "FFEF4444" for b in blocks)


def test_export_excel_highlights_judgment_reason_terms(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    resp_path = case_dir / "responses" / "JP2030000002A.json"
    response = json.loads(resp_path.read_text(encoding="utf-8"))
    response["comparisons"][0]["judgment_reason"] = "成分Aは記載されるが未知語は登録語ではない。"
    resp_path.write_text(json.dumps(response, ensure_ascii=False, indent=2), encoding="utf-8")

    result, code = export_excel("smoke")

    assert code == 200
    ws = load_workbook(result["path"], rich_text=True)["対比表"]
    value = ws["C10"].value
    assert isinstance(value, CellRichText)
    blocks = [b for b in value if isinstance(b, TextBlock)]
    assert any(b.text == "成分A" and b.font.color.rgb == "FFFF9999" for b in blocks)
    assert any(b.text == "成分A" and b.font.rFont == "メイリオ" for b in blocks)
    assert not any("未知語" in b.text and b.font.color.rgb == "FFEF4444" for b in blocks)


def test_export_excel_paste_sheet_uses_compact_notation(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    resp_path = case_dir / "responses" / "JP2030000002A.json"
    response = json.loads(resp_path.read_text(encoding="utf-8"))
    response["comparisons"][0]["judgment"] = "×"
    response["comparisons"][0]["cited_location"] = "10;/備考メモ;//防備録"
    resp_path.write_text(json.dumps(response, ensure_ascii=False, indent=2), encoding="utf-8")

    result, code = export_excel("smoke")

    assert code == 200
    ws = load_workbook(result["path"])["ペースト用"]
    assert ws["C3"].value == "!10/備考メモ"


def test_export_excel_paste_sheet_normalizes_same_kind_refs(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    resp_path = case_dir / "responses" / "JP2030000002A.json"
    response = json.loads(resp_path.read_text(encoding="utf-8"))
    response["comparisons"][0]["judgment"] = "○"
    response["comparisons"][0]["cited_location"] = "請求項1;CL2;表1;T2;【0001】/コメント"
    resp_path.write_text(json.dumps(response, ensure_ascii=False, indent=2), encoding="utf-8")

    result, code = export_excel("smoke")

    assert code == 200
    ws = load_workbook(result["path"])["ペースト用"]
    assert ws["C3"].value == "CL1,2;T1,2;1/コメント"


def test_export_full_report_hongan_analysis_strips_html(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    analysis = {
        "template_id": "hongan_v0.1",
        "version": "0.1",
        "sections": [{
            "id": 1,
            "title": "発明の本質",
            "items": [{
                "id": "1.1",
                "label": "発明の一言要約",
                "value": "<ul><li><<HL>>成分A<</HL>></li><li>請求項1</li></ul>",
            }],
        }],
    }
    (case_dir / "analysis").mkdir(exist_ok=True)
    (case_dir / "analysis" / "hongan_analysis.json").write_text(
        json.dumps(analysis, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    result, code = export_full_report("smoke")

    assert code == 200
    ws = load_workbook(result["path"])["本願解析結果"]
    assert ws["C3"].value == "・成分A\n・請求項1"


def test_export_full_report_highlights_inventive_step_terms(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    inv = {
        "overall_assessment": {
            "inventive_step": "なし",
            "reasoning": "成分Aと未知語を組み合わせても容易想到である。",
        },
        "primary_reference": {
            "document_id": "JP2030000002A",
            "selection_reason": "成分Bを含む主引用発明である。",
        },
    }
    (case_dir / "inventive_step.json").write_text(
        json.dumps(inv, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    result, code = export_full_report("smoke")

    assert code == 200
    ws = load_workbook(result["path"], rich_text=True)["進歩性判断"]
    rich_values = [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, CellRichText)
    ]
    assert rich_values
    blocks = [
        block
        for value in rich_values
        for block in value
        if isinstance(block, TextBlock)
    ]
    assert any(block.text == "成分A" and block.font.color.rgb == "FFFF9999" for block in blocks)
    assert any(block.text == "成分B" and block.font.color.rgb == "FFFF9999" for block in blocks)
    assert any(block.text == "成分A" and block.font.rFont == "メイリオ" for block in blocks)
    assert not any("未知語" in block.text and block.font.color.rgb == "FFEF4444" for block in blocks)


def test_export_full_report_uses_meiryo_and_valid_rich_text_colors(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    inv = {
        "overall_assessment": {"reasoning": "成分Aに基づき判断する。"},
        "primary_reference": {"selection_reason": "成分Bを含む。"},
    }
    (case_dir / "inventive_step.json").write_text(
        json.dumps(inv, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    result, code = export_full_report("smoke")

    assert code == 200
    wb = load_workbook(result["path"], rich_text=True)
    assert wb["対比表"]["A1"].font.name == "メイリオ"
    rich_blocks = [
        block
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, CellRichText)
        for block in cell.value
        if isinstance(block, TextBlock)
    ]
    assert rich_blocks
    assert all(block.font.rFont == "メイリオ" for block in rich_blocks if block.font)

    with zipfile.ZipFile(result["path"]) as z:
        xml = "\n".join(
            z.read(name).decode("utf-8")
            for name in z.namelist()
            if name.startswith("xl/worksheets/sheet")
        )
    assert not re.search(r'rgb="00[0-9A-Fa-f]{6}"', xml)


def test_check_segments_freshness_clean_fixture(copy_case_fixture):
    copy_case_fixture("smoke")

    result, code = check_segments_freshness("smoke")

    assert code == 200
    assert result["has_responses"] is True
    assert result["response_count"] == 1
    assert result["current_segment_count"] == 2
    assert result["missing_in_responses"] == []
    assert result["orphans_in_responses"] == {}
    assert result["needs_recompare"] is False


def test_check_segments_freshness_detects_missing_after_segment_added(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    segments = json.loads((case_dir / "segments.json").read_text(encoding="utf-8"))
    segments[0]["segments"].append({"id": "1C", "text": "成分Cを含む"})
    (case_dir / "segments.json").write_text(json.dumps(segments, ensure_ascii=False, indent=2), encoding="utf-8")

    result, code = check_segments_freshness("smoke")

    assert code == 200
    assert result["missing_in_responses"] == ["1C"]
    assert result["needs_recompare"] is True


def test_check_segments_freshness_detects_orphan_response_id(copy_case_fixture):
    case_dir = copy_case_fixture("smoke")
    response = json.loads((case_dir / "responses" / "JP2030000002A.json").read_text(encoding="utf-8"))
    response["comparisons"].append({
        "requirement_id": "9Z",
        "judgment": "×",
        "cited_location": "",
        "judgment_reason": "古い分節ID",
    })
    (case_dir / "responses" / "JP2030000002A.json").write_text(
        json.dumps(response, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    result, code = check_segments_freshness("smoke")

    assert code == 200
    assert result["orphans_in_responses"] == {"JP2030000002A": ["9Z"]}
    assert result["needs_recompare"] is True
